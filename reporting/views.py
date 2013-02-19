import datetime
import flask_login
import forms
import ldap
import logging
import json
import re
import tempfile
import time

from flask import render_template, request, redirect, url_for, session, abort, send_file
from flask.views import MethodView
from flask_login import login_required
from models import User, TimeRecord


class UserAwareView(MethodView):
    """
    A base view class to extend.
    """

    @property
    def session(self):
        """
        Adds the session property to the view.
        """
        return session

    @property
    def user(self):
        """
        Adds the user property to the view.

        :returns: The currently logged in user if one exists, else None
        """
        if not flask_login.current_user.is_anonymous():
            return flask_login.current_user._get_current_object()
        else:
            return None

    def get_context(self, extra_ctx=None, **kwargs):
        """
        Adds a helper function to the view to get the context.

        :returns: The current context with the user set.
        """
        if self.user and self.user.is_admin:
            ctx = {
                'user': self.user,
            }
            if extra_ctx:
                ctx.update(extra_ctx)
            ctx.update(kwargs)
            return ctx
        else:
            return {}


class Home(UserAwareView):
    """
    The view for the home page.
    """
    def get(self):
        context = self.get_context()
        
        if not self.user or not self.user.is_admin:
            context['form'] = forms.LoginForm()
            return render_template('index.html', **context)
        else:
            users = User.objects(is_team_admin=False)
            context = {
                'users': users,
                'user': self.user
            }
            return render_template('admin.html', **context)

class Login(UserAwareView):
    def post(self):
        form = forms.LoginForm(request.form)
        user = None

        username = form.username.data
        username = re.escape(username)
        password = form.password.data
        remember = form.remember_me.data

        if form.validate():
            try:
                logging.warning("Starting LDAP")
                conn = ldap.initialize('ldap://192.168.1.5')
                conn.protocol_version = 3
                conn.set_option(ldap.OPT_REFERRALS, 0)
                conn.simple_bind_s(username + '@site2.cdc.com', password)
            
                result_id = conn.search('DC=site2,DC=cdc,DC=com', ldap.SCOPE_SUBTREE, "(sAMAccountName=" + username + ")")
                                
                result_set = []
                while 1:
                    result_type, result_data = conn.result(result_id, 0)
                    if (result_data == []):
                        break
                    else:
                        if result_type == ldap.RES_SEARCH_ENTRY:
                            result_set.append(result_data)

                isAdmin = False
                isActive = False
                isApprover = False
                
                if len(result_set) > 0:
                    for i in range(len(result_set)):
                        for entry in result_set[i]:
                            entry_tuple = entry[1]
                            if len(entry_tuple) > 0:
                                for group in entry_tuple['memberOf']:
                                    logging.debug("Group: %s" % group)
                                    if "CN=www," in group:
                                        isActive = True
                                    if "CN=cdc-admins," in group:
                                        isAdmin = True
                                    if "CN=cdc-approvers," in group:
                                        isApprover = True
                
                # At this point, we have gotten the user from
                # the AD server, and verified that they are
                # active
                if isActive and isAdmin:
                    logging.debug("%s: isActive: %s" % (username, isActive))
                    logging.debug("%s: isAdmin: %s" % (username, isAdmin))
                    logging.debug("%s: isApprover: %s" % (username, isApprover))
                    user = User.get_user_by_username(username)
            
            except ldap.INVALID_CREDENTIALS:
                logging.warning("Invalid Credentials")
                user = None
            except ldap.SERVER_DOWN:
                logging.warning("Server down...")
                user = None
            
            if user:
                user.save()

                logging.debug("Authorized!")
                flask_login.login_user(user, remember=remember)
                return "success"
            
        return "Access Denied"


class Logout(UserAwareView):
    """
    The view for the logout page.
    """
    decorators = [login_required]

    def get(self):
        flask_login.logout_user()
        return redirect(url_for('home'))


class Export(UserAwareView):
    """
    The REST API endpoint for getting payroll info about a user.
    """
    decorators = [login_required]
    
    def get(self, username):
        if not self.user.is_admin:
            return redirect(url_for('home'))
        
        from openpyxl import Workbook        
        days = int(request.args.get('days', 14))

        user = User.get_user_by_username(username)
        if not user:
            abort(404)

        records = TimeRecord.get_approved_records_by_username(user.username, num_days=days)

        wb = Workbook()
        ws = wb.worksheets[0]
        ws.title = "Payroll Information"
    
        # User name
        ws.cell('%s%s' % ('A', 1)).value = 'User'
        ws.cell('%s%s' % ('A', 1)).style.font.bold = True
        ws.cell('%s%s' % ('B', 1)).value = user.username

        # SSN - TODO: should remove
        ws.cell('%s%s' % ('A', 2)).value = 'SSN'
        ws.cell('%s%s' % ('A', 2)).style.font.bold = True
        ws.cell('%s%s' % ('B', 2)).value = '***-**-%s' % user.ssn[-4:]
    
        # TimeRecord headers
        ws.cell('%s%s' % ('A', 4)).value = 'Date'
        ws.cell('%s%s' % ('A', 4)).style.font.bold = True
        ws.cell('%s%s' % ('B', 4)).value = 'Clock In'
        ws.cell('%s%s' % ('B', 4)).style.font.bold = True
        ws.cell('%s%s' % ('C', 4)).value = 'Clock Out'
        ws.cell('%s%s' % ('C', 4)).style.font.bold = True
        ws.cell('%s%s' % ('D', 4)).value = 'Approved?'
        ws.cell('%s%s' % ('D', 4)).style.font.bold = True
        ws.cell('%s%s' % ('E', 4)).value = 'Approved By'
        ws.cell('%s%s' % ('E', 4)).style.font.bold = True

        # All TimeRecords
        row = 5
        for record in records:
            ws.cell('%s%s' % ('A', row)).value = record.date.strftime('%B %d')
            ws.cell('%s%s' % ('B', row)).value = record.clock_in.strftime('%I:%M %p')
            ws.cell('%s%s' % ('C', row)).value = record.clock_out.strftime('%I:%M %p')
            ws.cell('%s%s' % ('D', row)).value = record.approved
            ws.cell('%s%s' % ('E', row)).value = record.approved_by
        
            row += 1

        filename = tempfile.mktemp()
        wb.save(filename=filename)

        return send_file(filename, 
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         attachment_filename="%s.xlsx" % user.username)
